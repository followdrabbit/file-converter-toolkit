# pdf_to_xlsx.py

# Importando as bibliotecas necessárias
import pdfplumber  # Para manipulação de arquivos PDF
from openpyxl import Workbook  # Para criar e trabalhar com arquivos Excel (XLSX)
import os  # Para operações relacionadas ao sistema de arquivos

# Definindo a classe PDFToXLSXConverter
class PDFToXLSXConverter:
    def __init__(self, pdf_path):
        """
        Construtor da classe PDFToXLSXConverter.

        :param pdf_path: Caminho para o arquivo PDF que será convertido.
        """
        self.pdf_path = pdf_path  # Armazena o caminho do arquivo PDF

    def convert(self):
        """
        Realiza a conversão do arquivo PDF para XLSX (Microsoft Excel).

        :return: O caminho do arquivo XLSX gerado.
        """
        wb = Workbook()  # Cria um novo documento Excel
        ws = wb.active  # Obtém a planilha ativa
        ws.title = "Page 1"  # Nomeia a primeira planilha

        try:
            # Abre o arquivo PDF para leitura
            with pdfplumber.open(self.pdf_path) as pdf:
                total_pages = len(pdf.pages)  # Calcula o total de páginas
                # Itera sobre cada página do PDF
                for i, page in enumerate(pdf.pages):
                    text = page.extract_text()  # Extrai o texto da página atual
                    if text:
                        # Adiciona cada linha do texto como uma nova linha na planilha
                        for j, line in enumerate(text.split('\n')):
                            ws.cell(row=j + 1, column=1, value=line)

                    # Se ainda houver páginas, cria uma nova planilha para a próxima
                    if i < total_pages - 1:
                        ws = wb.create_sheet(title=f"Page {i + 2}")
                    # Imprime o progresso da conversão
                    print(f"Processando página {i + 1}/{total_pages} para XLSX")
            
            # Define o nome do arquivo Excel baseado no nome do arquivo PDF
            xlsx_filename = os.path.splitext(self.pdf_path)[0] + '.xlsx'
            wb.save(xlsx_filename)  # Salva o arquivo Excel
            return xlsx_filename  # Retorna o caminho do arquivo Excel
        except Exception as e:
            # Captura e imprime qualquer erro que ocorra durante o processamento
            print(f"Erro ao processar o arquivo PDF: {e}")
            return None  # Retorna None em caso de erro

# Este bloco de código é executado se o script for o principal executado
if __name__ == "__main__":
    import sys  # Importa a biblioteca sys para manipulação de argumentos da linha de comando

    # Verifica se um caminho de arquivo foi fornecido como argumento
    if len(sys.argv) > 1:
        pdf_file = sys.argv[1]  # Armazena o caminho do arquivo PDF fornecido
        converter = PDFToXLSXConverter(pdf_file)  # Cria uma instância do conversor
        xlsx_file = converter.convert()  # Executa a conversão

        # Verifica se a conversão foi bem-sucedida
        if xlsx_file:
            print(f"Arquivo convertido e salvo como {xlsx_file}.")  # Informa o usuário sobre o sucesso da operação
        else:
            # Mensagem de erro se a conversão falhar
            print("Não foi possível converter o arquivo PDF.")
    else:
        # Mensagem de erro se nenhum arquivo for fornecido
        print("Por favor, forneça o caminho do arquivo PDF.")
